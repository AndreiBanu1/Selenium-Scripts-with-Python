import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

from WorkWithExcel import rows

options = Options()
options.add_argument("start-maximized")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.get("http://www.facebook.com/")

def getRowcount (file, sheetName) :
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def readData(file, sheetName, rownum, coloumnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return sheet.cell (row=rownum, column=coloumnno).value


path="E:\data\data.xlsx"
workbook=openpyxl.load_workbook(path)

rows-getRowcount (path, 'Sheet1')

for r in range(2,3):
    username=readData(path, "Sheetl", r, 1)
    password=readData (path, "Sheetl", r,2)
    print(username)
    print(password)
    driver.find_element_by_id("email").send_keys(username)
    driver.find_element_by_id("password").send_keys(password)
    driver.find_element_by_id("btn").click()